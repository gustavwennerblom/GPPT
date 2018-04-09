import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string


# noinspection PyPep8Naming
class ExcelParser:

    # Returns the version number (as string) of the GPPT client tool
    def determine_version(self):
        ws = self.wb.worksheets[4]
        version_string = ws['A1'].value
        # logging.info(version_string)
        version = version_string[-4:].strip()
        # logging.info("Found submission from tool with version %s" % version)
        print("Version: {}".format(version))
        return version

    # Support method to get the various alternative core sheets
    def get_sheet(self, sheetletter):
        try:
            version_number = float(self.determine_version())
        except ValueError as error:
            logging.error("Wrong format of tool version " + repr(error))
            version_number  = 0.4

        if sheetletter == "A":
            if version_number > 0.4:
                try:
                    return self.wb.get_sheet_by_name("A) Project pricing - consulting")
                    return self.wb.get_sheet_by_name("Project pricing - consulting")
                except KeyError as err:
                    logging.error("No pricing sheet found" + str(err))
                    return err
            else:
                return self.wb.get_sheet_by_name("Project pricing - consulting")
        # if sheetletter == "A":
        #     v04name = "Project pricing - consulting"
        #     v06name = "A) Project pricing consulting"
        #     if v04name in self.wb.get_sheet_names():
        #         return self.wb.get_sheet_by_name(v04name)
        #     elif v06name in self.wb.get_sheet_names():
        #         return self.wb.get_sheet_by_name(v06name)
        #     else:
        #         return self.wb.worksheets[1]

        elif sheetletter == "B":
            v02name = "Project planning"
            v04name = "B) Activity-role planning"
            if v02name in self.wb.get_sheet_names():
                return self.wb.get_sheet_by_name(v02name)
            elif v04name in self.wb.get_sheet_names():
                return self.wb.get_sheet_by_name(v04name)
            else:
                return self.wb.worksheets[2]
        elif sheetletter == "C":
            return self.wb.get_sheet_by_name("C) Activity-role-week planning")
        else:
            raise ExcelParsingError("Sheet names must be A, B, or C")

    # Returns lead office of the proposal
    def get_lead_office(self):
        try:
            ws = self.get_sheet("A")
        except KeyError as err:
            return "Unknown"
        return ws['C4'].value

    # Returns region
    def get_region(self):
        ws = self.get_sheet("A")
        return ws['N4'].value

    # Returns total project margin
    def get_margin(self):
        ws = self.get_sheet("A")
        if float(self.determine_version())>0.4:
            col=11
        else:
            col=13

        for cell in ws['B']:
            if cell.value == "SUBTOTAL":
                return ws.cell(row=cell.row, column=col).value
        raise ExcelParsingError("Cannot find project margin in project file")

    # Returns total project fee
    def get_project_fee(self):
        ws = self.get_sheet("A")
        for cell in ws['B']:
            if cell.value == "Project fee:":
                return ws.cell(row=cell.row, column=3).value
        raise ExcelParsingError("Cannot find project fee in project file")

    def get_blended_hourly_rate(self):
        ws = self.get_sheet("A")
        for cell in ws['B']:
            if cell.value == "Blended hourly rate:":
                return cell.offset(column=1).value
        raise ExcelParsingError("Cannot find blended hourly rate in project file")

    # Returns a dict of hours estimated by role
    def get_hours_by_role(self):
        ws = self.get_sheet("A")
        for cell in ws['B']:
            if cell.value == "Role":
                first_row = cell.row + 1
            elif cell.value == "SUBTOTAL":
                last_row = cell.row - 2

        roles_hours = {"Manager": 0, "SPM": 0,    "PM": 0, "Cons": 0, "Assoc": 0}
        for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=2, max_col=2):
            for cell in row:
                for key in roles_hours:
                    try:
                        if key in cell.value:
                            roles_hours[key] += int(cell.offset(column=8).value)
                    except ValueError as error:
                        logging.error(repr(error))
                        roles_hours[key] += -1     # Occurs in case of #REF error in excel sheet
                    except TypeError as error:
                        logging.error(repr(error))
                        roles_hours[key] += -1    # Saves from crash when returning None
        return roles_hours

    # Returns total number of hours estiamted
    def get_total_hours(self):
        ws = self.get_sheet("A")
        for cell in ws['13']:
            if cell.value == "Total hours used":
                total_hours_col = column_index_from_string(cell.column)
        for cell in ws['B']:
            if cell.value == "SUBTOTAL":
                total_hours_row = cell.row

        total_hours = ws.cell(row=total_hours_row, column=total_hours_col).value
        if total_hours == '#REF!':
            total_hours = 0

        return total_hours

    # Attempts to figure out which pricing method was used by the submitter
    def assess_pricing_method(self, db_id):

        try:
            ws_B = self.get_sheet("B")
            ws_C = self.get_sheet("C")
        except KeyError as error:
            logging.error("Erron or row %i " % db_id + repr(error))
            return "Auto-analysis failed, possibly too old tool version?"

        # Check total hours in main sheet
        hours_main = self.get_total_hours()

        # Check total hours in sheet B
        for t in tuple(ws_B.rows):
            for cell in t:
                if cell.value == "TOTAL BY ACTIVITY":
                    target_B_col = column_index_from_string(cell.column)
                if cell.value == "TOTAL HOURS BY ROLE":
                    target_B_row = cell.row

        hours_B = ws_B.cell(row=target_B_row, column=target_B_col).value

        # Check total hours in sheet C
        for t in tuple(ws_C.rows):
            for cell in t:
                if cell.value == "TOTAL BY ACTIVITY":
                    target_C_col = column_index_from_string(cell.column)
                if cell.value == "TOTAL HOURS BY WEEK":
                    target_C_row = cell.row

        hours_C = ws_C.cell(row=target_C_row, column=target_C_col).value

        # Check if week-share method has been used
        # TO BE CONSTRUCTED

        # Final check for which method has been used and return string of that result
        try:
            if int(hours_B) in range(int(hours_main*0.95), int(hours_main*1.05)):
                return "Method 3 (Activity-Role)"
            elif int(hours_C) in range(int(hours_main*0.95), int(hours_main*1.05)):
                return "Method 4 (Activity-Role-Week)"
            else:
                return "Method 1 or 2"
        except TypeError as error:
            logging.error(repr(error))
            return "Method unknown (auto-analysis failed)"

    # Initializing with reference to a filename
    def __init__(self, tempfile):
        self.wb = Workbook()
        self.wb = load_workbook(tempfile, data_only=True)
        self.filename = tempfile


class ExcelParsingError(Exception):
    pass
